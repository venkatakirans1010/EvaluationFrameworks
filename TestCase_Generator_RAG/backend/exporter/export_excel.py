"""
Excel exporter module.
Exports generated test cases to Excel format with proper formatting.
"""

import re
from pathlib import Path
from typing import List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Required columns based on the test case specification
REQUIRED_COLUMNS = [
    "Test ID",
    "Test Title",
    "Preconditions",
    "Steps",
    "Expected Result",
    "Type",
    "Priority",
    "Linked Requirements"
]


def _parse_markdown_table(markdown_table: str) -> List[List[str]]:
    """
    Parse a Markdown table string into a list of rows (each row is a list of cells).
    
    Args:
        markdown_table: Markdown table string with pipe separators
        
    Returns:
        List of rows, where each row is a list of cell strings
    """
    lines = markdown_table.strip().split('\n')
    rows = []
    
    for line in lines:
        # Skip empty lines
        if not line.strip():
            continue
        
        # Skip separator lines (e.g., |---|---|)
        if re.match(r'^\s*\|[\s\-:]+\|', line):
            continue
        
        # Parse row: split by pipe and clean up
        # Remove leading/trailing pipes if present
        line = line.strip()
        if line.startswith('|'):
            line = line[1:]
        if line.endswith('|'):
            line = line[:-1]
        
        # Split by pipe and strip whitespace from each cell
        cells = [cell.strip() for cell in line.split('|')]
        rows.append(cells)
    
    return rows


def _auto_fit_column_widths(worksheet, df: pd.DataFrame):
    """
    Auto-fit column widths in Excel worksheet.
    
    Args:
        worksheet: openpyxl worksheet object
        df: DataFrame to get column names from
    """
    for idx, column in enumerate(df.columns, start=1):
        column_letter = get_column_letter(idx)
        
        # Calculate max width needed
        max_length = len(str(column))
        
        # Check all rows in this column
        for row_idx, cell_value in enumerate(df[column].astype(str), start=2):
            # Account for multi-line cells
            cell_lines = str(cell_value).split('\n')
            max_cell_length = max(len(line) for line in cell_lines)
            max_length = max(max_length, max_cell_length)
        
        # Set column width (add some padding)
        # Cap at 50 characters to prevent extremely wide columns
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def export_markdown_table_to_excel(
    markdown_table: str, 
    output_path: str = "./generated_ui_testcases.xlsx"
) -> str:
    """
    Export a Markdown table to Excel format.
    
    Args:
        markdown_table: Markdown table string (may include surrounding text)
        output_path: Path where Excel file should be saved
        
    Returns:
        str: Absolute path to the saved Excel file
        
    Raises:
        ValueError: If required columns are missing or table cannot be parsed
    """
    # Extract table from markdown text (handle cases where there's text before/after)
    # Look for table pattern: lines starting with | or containing multiple pipes
    lines = markdown_table.split('\n')
    table_lines = []
    in_table = False
    
    for line in lines:
        # Check if line looks like a table row (contains multiple pipes)
        if '|' in line and line.count('|') >= 2:
            in_table = True
            table_lines.append(line)
        elif in_table and line.strip() and '|' not in line:
            # End of table
            break
    
    if not table_lines:
        raise ValueError("No Markdown table found in the provided text. Expected a table with pipe separators.")
    
    markdown_table_text = '\n'.join(table_lines)
    
    # Parse Markdown table
    rows = _parse_markdown_table(markdown_table_text)
    
    if len(rows) < 2:
        raise ValueError("Markdown table must have at least a header row and one data row.")
    
    # Extract header and data rows
    header_row = rows[0]
    data_rows = rows[1:]
    
    # Normalize header names (strip whitespace, handle case variations)
    normalized_header = [col.strip() for col in header_row]
    
    # Validate required columns exist
    normalized_actual = [col.lower() for col in normalized_header]
    
    missing_columns = []
    for req_col in REQUIRED_COLUMNS:
        req_col_lower = req_col.lower()
        if not any(req_col_lower in act_col or act_col in req_col_lower for act_col in normalized_actual):
            missing_columns.append(req_col)
    
    if missing_columns:
        raise ValueError(
            f"Missing required columns: {', '.join(missing_columns)}. "
            f"Found columns: {', '.join(normalized_header)}. "
            f"Required columns: {', '.join(REQUIRED_COLUMNS)}"
        )
    
    # Create DataFrame
    # Pad data rows to match header length if needed
    max_cols = len(normalized_header)
    padded_data_rows = []
    for row in data_rows:
        padded_row = row + [''] * (max_cols - len(row))
        padded_data_rows.append(padded_row[:max_cols])
    
    df = pd.DataFrame(padded_data_rows, columns=normalized_header)
    
    # Remove empty rows
    df = df.dropna(how='all')
    
    # Ensure output directory exists
    output_path_obj = Path(output_path)
    output_path_obj.parent.mkdir(parents=True, exist_ok=True)
    
    # Write to Excel
    df.to_excel(output_path, index=False, engine='openpyxl')
    
    # Auto-fit column widths using openpyxl
    workbook = load_workbook(output_path)
    worksheet = workbook.active
    _auto_fit_column_widths(worksheet, df)
    workbook.save(output_path)
    
    # Return absolute path
    absolute_path = output_path_obj.resolve()
    return str(absolute_path)

